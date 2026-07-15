#!/usr/bin/env python3
"""Build the week-by-week tracker .xlsx from tracker_data.json."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "/Users/fmy-235/Desktop/Training-Tracker"
FUTURE_WEEKS = 6  # blank week columns to fill in going forward

data = json.load(open(f"{OUT_DIR}/tracker_data.json"))

FONT = "Arial"
navy = "1F3A5F"; gold = "D4A043"; lightgold = "FBF1DD"; grey = "F2F2F2"
white = "FFFFFF"
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(c):
    c.font = Font(name=FONT, bold=True, color=white, size=11)
    c.fill = PatternFill("solid", fgColor=navy)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

def base(c, bold=False, size=10, fill=None, align="left", color="222222", italic=False):
    c.font = Font(name=FONT, bold=bold, size=size, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border

wb = Workbook()

# ============================================================ How to use
ws = wb.active
ws.title = "How to use"
ws.sheet_view.showGridLines = False
ws["B2"] = "Training Tracker — GPP S&C programme"
ws["B2"].font = Font(name=FONT, bold=True, size=18, color=navy)
ws["B3"] = "Programme start: 28 Jun 2026  ·  built from RepCount export"
ws["B3"].font = Font(name=FONT, size=11, color="666666", italic=True)

rows = [
    ("", ""),
    ("HOW THIS WORKS", ""),
    ("Each session tab", "One row per exercise, grouped into its superset block. One column per week — the cell shows every set as weight×reps (BW = bodyweight)."),
    ("Superset blocks", "Exercises in the same block (Prep / Block 1 / 2 / 3…) are performed back-to-back as a superset. RepCount doesn't store this — it's reconstructed from the trainer's programme."),
    ("Reading progress", "Scan a row left→right: rising weight or reps = progression. Same = held. The dashboard turns each row into a trend line."),
    ("", ""),
    ("UPDATING EACH WEEK", ""),
    ("Option A — paste", "Paste the RepCount session text into chat; I'll add the new week column and refresh the dashboard."),
    ("Option B — CSV", "Drop a fresh RepCount CSV export; I'll parse the new sessions automatically."),
    ("By hand", "Type sets straight into the next empty W-column (format: 20×8 · 22.5×8). Blank W3–W8 columns are ready."),
    ("", ""),
    ("TABS", ""),
    ("GPP S&C 1 / 2 / 3", "The three strength sessions. S&C 3 is a placeholder until its first session is logged."),
    ("Runs", "Two runs per week — one Sprint session, one Long/slow run. Pace auto-calculates from distance + time."),
    ("", ""),
    ("NOTES ON THE DATA", ""),
    ("Cleaned", "'1-arm 3-point row 12.5×80' on 28 Jun was a typo → corrected to ×10. Exercise-name spellings normalised so weeks match."),
]
r = 5
for a, b in rows:
    ca, cb = ws.cell(r, 2, a), ws.cell(r, 3, b)
    if a and not b:  # section header
        ca.font = Font(name=FONT, bold=True, size=12, color=gold)
    else:
        ca.font = Font(name=FONT, bold=True, size=10, color=navy)
        cb.font = Font(name=FONT, size=10, color="333333")
        cb.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 88

# ============================================================ session tabs
def build_session(name, sess, has_data=True):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = name
    ws["A1"].font = Font(name=FONT, bold=True, size=16, color=navy)
    weeks = sess["weeks"]
    logged = [f"W{w['n']}\n{w['date'][8:10]}/{w['date'][5:7]}" for w in weeks]
    # header at row 3
    headers = ["Block", "Exercise"] + logged
    # future blank weeks
    start_future = len(weeks) + 1
    for i in range(FUTURE_WEEKS):
        headers.append(f"W{start_future + i}\n(date)")
    headers.append("Notes")
    hr = 3
    for j, h in enumerate(headers, start=1):
        c = ws.cell(hr, j, h)
        style_header(c)
    # body
    row = hr + 1
    shade = False
    for b in sess["blocks"]:
        block_start = row
        exs = b["exercises"]
        for k, e in enumerate(exs):
            fill = lightgold if shade else white
            ws.cell(row, 1, b["label"] if k == 0 else "")
            base(ws.cell(row, 1), bold=True, fill=fill, align="center", color=navy)
            base(ws.cell(row, 2, e["name"]), bold=True, fill=fill)
            col = 3
            for wk in e["byweek"]:
                base(ws.cell(row, col, wk["cell"]), fill=fill, align="center")
                col += 1
            for _ in range(FUTURE_WEEKS):
                base(ws.cell(row, col, ""), fill=fill, align="center")
                col += 1
            base(ws.cell(row, col, ""), fill=fill)  # notes
            row += 1
        # merge block label cell
        if len(exs) > 1:
            ws.merge_cells(start_row=block_start, start_column=1, end_row=row - 1, end_column=1)
        shade = not shade
    # widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 24
    for j in range(3, 3 + len(weeks) + FUTURE_WEEKS):
        ws.column_dimensions[get_column_letter(j)].width = 15
    ws.column_dimensions[get_column_letter(3 + len(weeks) + FUTURE_WEEKS)].width = 26
    ws.freeze_panes = "C4"
    ws.row_dimensions[hr].height = 30
    if not has_data:
        ws.cell(2, 1, "Placeholder — paste the first GPP S&C 3 session and I'll populate it.").font = \
            Font(name=FONT, italic=True, size=11, color="B00000")

for name, sess in data["sessions"].items():
    build_session(name, sess)

# GPP S&C 3 placeholder (no structure yet)
ws3 = wb.create_sheet("GPP S&C 3")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "GPP S&C 3"
ws3["A1"].font = Font(name=FONT, bold=True, size=16, color=navy)
ws3["A3"] = "Not logged yet."
ws3["A3"].font = Font(name=FONT, italic=True, size=12, color="B00000")
ws3["A4"] = "Paste the first GPP S&C 3 session (with A/B/C superset letters) and this tab fills in like the others."
ws3["A4"].font = Font(name=FONT, size=11, color="333333")
ws3.column_dimensions["A"].width = 90

# ============================================================ Runs tab
ws = wb.create_sheet("Runs")
ws.sheet_view.showGridLines = False
ws["A1"] = "Runs — 2 per week"
ws["A1"].font = Font(name=FONT, bold=True, size=16, color=navy)
ws["A2"] = "1 × Sprint session  ·  1 × Long / slow-distance run"
ws["A2"].font = Font(name=FONT, size=11, color="666666", italic=True)
headers = ["Week", "Date", "Type", "Distance (km)", "Time (min)", "Pace (min/km)", "RPE", "Notes"]
hr = 4
for j, h in enumerate(headers, 1):
    style_header(ws.cell(hr, j, h))
# example row + blank template rows
ws.cell(hr + 1, 1, "e.g. W1")
ws.cell(hr + 1, 2, "20/07")
ws.cell(hr + 1, 3, "Sprint")
ws.cell(hr + 1, 4, 3.0)
ws.cell(hr + 1, 5, 18)
ws.cell(hr + 1, 7, 8)
ws.cell(hr + 1, 8, "example — 6×200m")
for j in range(1, 9):
    base(ws.cell(hr + 1, j), fill=grey, italic=True, align="center" if j not in (8,) else "left")
ws.cell(hr + 1, 6, "=IF(D5=0,\"\",E5/D5)")
base(ws.cell(hr + 1, 6), fill=grey, italic=True, align="center")

nrows = 24  # 12 weeks × 2 runs
for i in range(nrows):
    r = hr + 2 + i
    typ = "Sprint" if i % 2 == 0 else "Long"
    wk = f"W{i//2 + 1}"
    ws.cell(r, 1, wk if i % 2 == 0 else "")
    ws.cell(r, 3, typ)
    ws.cell(r, 6, f"=IF(D{r}=0,\"\",E{r}/D{r})")
    fill = white if (i // 2) % 2 == 0 else lightgold
    for j in range(1, 9):
        base(ws.cell(r, j), fill=fill, align="center" if j != 8 else "left")
    ws.cell(r, 6).number_format = "0.00"
ws.cell(hr + 1, 6).number_format = "0.00"
widths = [8, 10, 10, 14, 12, 14, 7, 34]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A5"
ws.row_dimensions[hr].height = 28

path = f"{OUT_DIR}/Training-Tracker.xlsx"
wb.save(path)
print("saved", path)
